# -*- encoding=utf8 -*-
__author__ = "Issac"

from os.path import abspath, dirname

import allure
import cv2
import io
import json
import jsonpath
import openpyxl
import pyautogui
import sys
from airtest.core.api import *
from airtest_selenium.proxy import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

project_path = dirname(dirname(abspath(__file__)))


class web_tool:
    def __init__(self, driver=None, json_data=None):
        self.driver = driver
        self.json_data = json_data

    def find_element_by_xpath_click(self, xpath):
        with allure.step("Click content："+xpath):
            self.driver.implicitly_wait(5)
            # WebDriverWait(self.driver,40).until()
            element = self.driver.find_element_by_xpath(xpath)
            ActionChains(self.driver).move_to_element(element).perform()
            sleep(1)
            ActionChains(self.driver).click(element).perform()

    def find_element_by_xpath_set_option(self, xpath, option):
        with allure.step(f"Select content: {xpath} and set option = {option}"):
            self.driver.implicitly_wait(20)
            self.driver.find_element_by_xpath(f"{xpath}/option[text()='{option}']").click()

    def find_elements_by_xpath(self, xpath):
        return self.driver.find_elements(By.XPATH, xpath)


    def find_element_by_xpath(self, xpath):
        element = self.driver.find_element(By.XPATH, xpath)
        return element

    def get_attribute_value_by_xpath(self, xpath):
        value = self.driver.find_element_by_xpath(By.XPATH, xpath).get_attribute('value')
        return value

    def execute_script(self, script):
        self.driver.execute_script(script)

    def get_current_window_handle(self):
        return self.driver.current_window_handle

    def switch_window(self, window_to_be_switched):
        switch_object = self.driver.switch_to
        return switch_object.window(window_to_be_switched)

    def switch_frame(self, frame_to_be_switched):
        switch_object = self.driver.switch_to
        return switch_object.frame(frame_to_be_switched)

    def switch_to_original_frame(self):
        switch_object = self.driver.switch_to
        return switch_object.defaultContent()


    def refresh(self):
        self.driver.refresh()

    # 断言元素是否存在
    def assert_element(self, xpath, msg=None):
        with allure.step("Assert content：" + xpath):
            self.driver.assert_exist(xpath, "xpath", msg)

    def get_element_text_content_by_xpath(self, xpath):
        return self.driver.find_element_by_xpath(xpath).text
    
    def find_element_by_xpath_send_keys(self, xpath, content):
        with allure.step("Input text：" + content):
            self.driver.find_element_by_xpath(xpath).click()
            self.driver.find_element_by_xpath(xpath).send_keys(Keys.CONTROL, 'a')
            self.driver.find_element_by_xpath(xpath).send_keys(Keys.BACKSPACE)
            self.driver.find_element_by_xpath(xpath).send_keys(content)

    def find_element_by_xpath_send_keys_without_clearing(self, xpath, content):
        with allure.step("Input text (without removing): " + content):
            self.driver.find_element_by_xpath(xpath).send_keys(content)

    def find_element_by_xpath_upload_file(self, xpath, file_url):
        self.driver.find_element_by_xpath(xpath).send_keys(file_url)

    # 物理输入
    def typewrite(self, content):
        pyautogui.typewrite(content)

    # 滑动到元素可见
    def scrollIntoView(self, xpath):
        self.driver.implicitly_wait(30)
        element = self.driver.find_element_by_xpath(xpath)
        self.driver.execute_script("arguments[0].scrollIntoView();", element)

    # 读取 Excel 文件并获取工作表
    def excel_data(self, file_path):
        params = []
        workbook = openpyxl.load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = []
                for cell in row:
                    row_data.append(cell)
                params.append(row_data)  # 将行数据添加到参数列表中
        return params

    # device,需要切换的设备 Windows/android
    def connect_driver(self, device=None):
        if not G.DEVICE_LIST:
            dev1 = connect_device("Android:///")
            dev2 = connect_device("Windows:///")
            # G.DEVICE_LIST.extend([dev1, dev2])
        print("当前连接设备如下")
        print(G.DEVICE_LIST)
        # 如果传入了 device 参数，则在设备列表中查找并返回对应设备
        if device is not None:
            if device == "Windows":
                print("切换Windows设备")
                set_current(1)
            elif device == "android":
                print("切换android设备")
                set_current(0)
            else:
                print("输入切换设备参数有误")

    # 根据jsonpath获取对应json值
    def get_jsonpath_data(self, path):
        data = json.loads(self.json_data)
        result = jsonpath.jsonpath(data, path)[0]
        return result
    
    # 截图
    def get_screenshot(self):
        sleep(2)
        self.driver.get_screenshot_as_file(f"{project_path}//data//code.png")

    # 二维码解码
    def decode(self):
        qrcode_filename = f'{project_path}//data//code.png'
        qrcode_image = cv2.imread(qrcode_filename)
        qrcode_detector = cv2.QRCodeDetector()
        data, b, c = qrcode_detector.detectAndDecode(qrcode_image)
        # print(b)
        # print(c)
        # print(data)
        return data
